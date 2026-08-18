#!/usr/bin/env python3
"""Update MAFATEEH closure workbook (sheets 2-9)."""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "MAFATEEH_RECONCILED_MASTER11-8-2026.xlsx"

def decision_cell(d):
    opts = ['APPROVE', 'MODIFY', 'REJECT', 'FURTHER INVESTIGATION']
    return '  '.join([f'{"☑" if o == d else "☐"} {o}' for o in opts])

# Comprehensive issue data
ISSUES = {
    'MAF-TECH-001': {
        'decision': 'APPROVE',
        'dev_comment': 'Implemented a real robots.txt file served as plain text via nginx, replacing the previous SPA HTML response. [This tells Google which pages it may crawl — like a “welcome” sign for search engines instead of a blank page.]',
        'next_action': 'Verify on production after deploy (curl /robots.txt).',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Added website/public/robots.txt and nginx exact location block. Completed Aug 2026. [Search engines now receive proper crawl instructions.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'File exists in repo; nginx serves text/plain. Post-deploy live check recommended. [Confirmed in code; quick live test after release.]',
        'after_evidence': 'public/robots.txt + nginx.default.conf location = /robots.txt',
    },
    'MAF-TECH-003': {
        'decision': 'APPROVE',
        'dev_comment': 'Canonical link tags are injected per route for both /en and /ar pages, with a static fallback in index.html. [Each page now tells Google its “official” URL so duplicate-language versions are not confused.]',
        'next_action': 'Spot-check canonical on key pages after deploy.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'SeoProvider + static canonical in index.html. [Official page addresses are set automatically.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Canonical logic confirmed in SeoProvider. [Built into the site; verify on live URLs after deploy.]',
        'after_evidence': 'SeoProvider.jsx + index.html canonical',
    },
    'MAF-TECH-012': {
        'decision': 'MODIFY',
        'dev_comment': 'We are keeping React (no SSR migration). Mitigated with static meta tags, JSON-LD in HTML, bilingual URLs, sitemap, and robots rules. Full page HTML without JavaScript is not in scope. [The site still loads content after scripts run, but we added the main SEO signals Google needs without rebuilding the whole site.]',
        'next_action': 'Monitor Google Search Console rendering after deploy; no SSR planned.',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'CSR architecture retained by business decision. Partial SEO mitigations applied. [We improved search visibility within the current website technology.]',
        'verify_status': 'PARTIALLY VERIFIED', 'result': 'PARTIAL',
        'verify_notes': 'Mitigations in code; full CSR limitation accepted. [Not a full fix, but acceptable given we stay on React.]',
        'after_evidence': 'Static meta/JSON-LD + bilingual routes',
    },
    'MAF-TECH-027': {
        'decision': 'APPROVE',
        'dev_comment': 'Structured data (JSON-LD) added: Organization in raw HTML; after load adds WebSite, BreadcrumbList, and BlogPosting on blog pages. [Rich search snippets — company info and breadcrumbs — can now appear in Google results.]',
        'next_action': 'Validate with Google Rich Results Test after deploy.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'index.html Organization schema + SeoProvider dynamic JSON-LD. [Search engines get structured company and page data.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'JSON-LD present in HTML and SeoProvider. [Ready for Google’s structured-data checker after go-live.]',
        'after_evidence': 'index.html + SeoProvider JSON-LD blocks',
    },
    'MAF-TECH-036': {
        'decision': 'APPROVE',
        'dev_comment': 'Dedicated /en/... and /ar/... URL routes implemented; language toggle updates the URL. hreflang tags and bilingual sitemap included. [Arabic and English now have separate web addresses Google can index independently.]',
        'next_action': 'Confirm hreflang in Search Console after deploy.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Locale routing, redirects from legacy URLs, sitemap + hreflang. [Arabic content is now discoverable via its own links.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Routes, sitemap, and hreflang confirmed in repo. [Both languages have their own URLs in the sitemap.]',
        'after_evidence': '/en/* /ar/* routes + sitemap hreflang',
    },
    'MAF-TECH-002': {
        'decision': 'APPROVE',
        'dev_comment': 'XML sitemap created with all main pages in English and Arabic, linked from robots.txt. [A map of all site pages is available for Google to find and index faster.]',
        'next_action': 'Submit sitemap in Google Search Console (Marketing).',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'website/public/sitemap.xml with en/ar URLs. [Complete page list for search engines.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'sitemap.xml in repo with bilingual entries. [File ready; Marketing to submit in GSC.]',
        'after_evidence': 'public/sitemap.xml',
    },
    'MAF-TECH-004': {
        'decision': 'FURTHER INVESTIGATION',
        'dev_comment': 'Index status cannot be confirmed without Google Search Console access. This is an owner/marketing account task, not a development code fix. [We need login access to Google’s search dashboard to see if pages are actually appearing in search results.]',
        'next_action': 'Marketing: grant GSC access and submit sitemap.',
        'impl_status': 'PENDING', 'impl_owner': 'Marketing',
        'impl_notes': 'Sitemap and SEO fixes are ready; verification blocked on GSC access. [Dev work done; marketing must check Google’s side.]',
        'verify_status': 'PENDING EXTERNAL', 'result': 'PENDING',
        'verify_notes': 'Requires Google Search Console — no IT Ops team; Marketing to action. [Someone with Google account access must complete this step.]',
        'after_evidence': 'Awaiting GSC access',
    },
    'MAF-TECH-005': {
        'decision': 'MODIFY',
        'dev_comment': 'Probe paths (wp-admin, etc.) return real HTTP 404 via nginx; unknown SPA routes get noindex meta. True HTTP 404 for every invalid React route would require SSR. [Fake “success” pages for hacker probes are blocked; remaining soft-404s are mitigated with “do not index” tags.]',
        'next_action': 'None — accepted partial mitigation.',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'nginx 404 for probes + SeoProvider noindex on unknown paths. [Better than before; full fix not possible without SSR.]',
        'verify_status': 'PARTIALLY VERIFIED', 'result': 'PARTIAL',
        'verify_notes': 'Probe 404s and noindex logic in code. [Security probe paths fixed; some invalid URLs still return 200 by design.]',
        'after_evidence': 'nginx probe blocks + SeoProvider noindex',
    },
    'MAF-TECH-006': {
        'decision': 'APPROVE',
        'dev_comment': 'HTTP to HTTPS and non-www to www redirects configured in deployment nginx. [Visitors are automatically sent to the secure, correct website address.]',
        'next_action': 'Verify redirects live after deploy.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Deploy nginx redirect rules. [All traffic goes to https://www.mafateehgroup.com.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Redirect config in deploy nginx. [Standard security redirect — confirm with browser after deploy.]',
        'after_evidence': 'deploy nginx HTTPS/www redirects',
    },
    'MAF-TECH-013': {
        'decision': 'APPROVE',
        'dev_comment': 'Route-level code splitting implemented via Vite manualChunks and React.lazy for non-home pages. [The site loads smaller pieces of code per page so pages open faster.]',
        'next_action': 'Review bundle report after build if needed.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'vite.config.js manualChunks + lazy routes in App.jsx. [Each page downloads only what it needs.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Build produces separate chunks. [Confirmed in build output.]',
        'after_evidence': 'Vite chunks + React.lazy routes',
    },
    'MAF-TECH-017': {
        'decision': 'APPROVE',
        'dev_comment': 'Open Graph and Twitter Card meta tags added statically in index.html and updated dynamically per route via SeoProvider. [Links shared on social media now show proper title, description, and image previews.]',
        'next_action': 'Test share preview on LinkedIn/Facebook after deploy.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Static OG in index.html + dynamic SeoProvider tags. [Social sharing previews are configured.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'OG/Twitter tags in HTML and SeoProvider. [Use Facebook/LinkedIn debug tools after deploy.]',
        'after_evidence': 'index.html + SeoProvider OG/Twitter',
    },
    'MAF-TECH-018': {
        'decision': 'MODIFY',
        'dev_comment': 'Bundle size reduced through code splitting and vendor chunking. Remaining size is expected for a React SPA with animations. [The download is smaller than before, but a rich interactive site will always be heavier than a plain HTML page.]',
        'next_action': 'Monitor Lighthouse after deploy.',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Splitting applied; full sub-500KB target unrealistic for current feature set. [Improved speed without removing features.]',
        'verify_status': 'PARTIALLY VERIFIED', 'result': 'PARTIAL',
        'verify_notes': 'Chunks reduced vs monolith; measure post-deploy. [Better, not perfect — normal for React sites.]',
        'after_evidence': 'Vite build chunk sizes',
    },
    'MAF-TECH-019': {
        'decision': 'APPROVE',
        'dev_comment': 'Browser and CDN cache headers configured: hashed assets cached 1 year; static SEO files have appropriate TTL; index.html set to no-store. [Repeat visitors load the site faster because their browser keeps copies of images and scripts.]',
        'next_action': 'Confirm Cache-Control headers on live site.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'nginx Cache-Control + CDN-Cache-Control headers. [Files are stored locally on repeat visits.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Caching rules in nginx.default.conf. [Check response headers after deploy.]',
        'after_evidence': 'nginx caching headers',
    },
    'MAF-TECH-022': {
        'decision': 'MODIFY',
        'dev_comment': 'Lab-estimated LCP improved via code splitting, font preload/swap, WebP images, and animation optimization (including Mac performance fix). Full LCP target may still need SSR for first paint. [Main content should appear faster; exact speed should be measured on the live site.]',
        'next_action': 'Run PageSpeed Insights after deploy.',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Performance optimizations applied Aug 2026 incl. ParticleField and WebP. [Page load speed improved through multiple fixes.]',
        'verify_status': 'NOT VERIFIED', 'result': 'PENDING MEASUREMENT',
        'verify_notes': 'Lab estimate only — needs Lighthouse/PageSpeed on production. [Dev improvements done; marketing can run speed test after launch.]',
        'after_evidence': 'Splitting, WebP, font preload, animation tuning',
    },
    'MAF-TECH-023': {
        'decision': 'MODIFY',
        'dev_comment': 'Total Blocking Time improved via code splitting and reduced animation cost. React execution remains on the main thread by architecture. [The site freezes less on load, but some delay is normal for interactive React apps.]',
        'next_action': 'Measure TBT with Lighthouse after deploy.',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Code splitting + animation performance fixes. [Smoother interaction during page load.]',
        'verify_status': 'NOT VERIFIED', 'result': 'PENDING MEASUREMENT',
        'verify_notes': 'Lab estimate — verify with Lighthouse post-deploy. [Needs a speed test on the live website.]',
        'after_evidence': 'Code splitting + reduced GPU load',
    },
    'MAF-TECH-038': {
        'decision': 'APPROVE',
        'dev_comment': 'html lang and dir attributes update dynamically based on locale (/en or /ar) and localStorage before first paint. [The browser knows whether the page is English or Arabic for accessibility and search.]',
        'next_action': 'None.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'index.html inline script + SeoProvider locale sync. [Language direction set correctly for each version.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'lang/dir logic in index.html and LanguageProvider. [Arabic pages show dir=rtl.]',
        'after_evidence': 'index.html + locale routing',
    },
    'MAF-TECH-049': {
        'decision': 'APPROVE',
        'dev_comment': 'Security headers added via nginx: X-Content-Type-Options, X-Frame-Options, Referrer-Policy, Permissions-Policy, and HSTS (also covers MAF-TECH-048). [Standard website security settings that protect visitors and improve trust signals.]',
        'next_action': 'Verify headers with securityheaders.com after deploy.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'nginx add_header directives on website container. [Security headers active on all responses.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Headers defined in nginx.default.conf. [Can be checked with free online header tools.]',
        'after_evidence': 'nginx security headers incl. HSTS',
    },
    'MAF-TECH-008': {
        'decision': 'MODIFY',
        'dev_comment': 'Crawl-trap risk mitigated same as MAF-TECH-005: probe paths return 404; unknown SPA pages get noindex. Full elimination of infinite SPA URLs needs SSR. [Search engines are steered away from junk paths; complete fix limited by React architecture.]',
        'next_action': 'None — accepted partial mitigation.',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Combined with soft-404 mitigation. [Reduces risk of Google wasting crawl budget.]',
        'verify_status': 'PARTIALLY VERIFIED', 'result': 'PARTIAL',
        'verify_notes': 'Probe blocks + noindex in place. [Partial protection in place.]',
        'after_evidence': 'Same as MAF-TECH-005',
    },
    'MAF-TECH-009': {
        'decision': 'APPROVE',
        'dev_comment': '/blog permanently redirects to /en/blogs via nginx 301. [Old blog link bookmarks and external links still work.]',
        'next_action': 'None.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'nginx location = /blog return 301. [Legacy URL redirected correctly.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Redirect in nginx.default.conf. [Old /blog address forwards to blogs page.]',
        'after_evidence': 'nginx /blog → /en/blogs',
    },
    'MAF-TECH-011': {
        'decision': 'APPROVE',
        'dev_comment': 'Blog post routes verified at /en/blogs/:slug and /ar/blogs/:slug. Missing slugs redirect to the blogs listing. [Individual blog articles have working addresses in both languages.]',
        'next_action': 'Marketing: publish test blog post and confirm URL.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'BlogPost.jsx + locale routes + redirect on 404. [Blog URLs work and handle missing posts gracefully.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Routes and error handling confirmed. [Blog links work; test with a real post after deploy.]',
        'after_evidence': 'App.jsx blog routes + BlogPost redirect',
    },
    'MAF-TECH-020': {
        'decision': 'FURTHER INVESTIGATION',
        'dev_comment': 'Origin cache headers (Cache-Control + CDN-Cache-Control) are implemented in nginx. Cloudflare dashboard cache rules still need to be applied — IT Dev prepared guide at deploy/cloudflare/CACHE_RULES.md (no separate Ops team). [The server is ready to cache; someone must click 3 rules in the Cloudflare website panel — about 5 minutes.]',
        'next_action': 'IT Dev or Marketing admin: apply 3 Cloudflare cache rules from CACHE_RULES.md.',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Origin headers done; Cloudflare UI rules pending (~5 min). [Half done in code; final step is a Cloudflare login task.]',
        'verify_status': 'PENDING EXTERNAL', 'result': 'PENDING',
        'verify_notes': 'Origin ready; Cloudflare dashboard config outstanding. [Dev team finished their part; Cloudflare account holder must finish.]',
        'after_evidence': 'nginx CDN-Cache-Control + CACHE_RULES.md guide',
    },
    'MAF-TECH-021': {
        'decision': 'APPROVE',
        'dev_comment': 'Vite produces content-hashed asset filenames with max-age=31536000 immutable cache headers. [Updated files get new names so browsers always fetch fresh code after a release.]',
        'next_action': 'None.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Vite hashing + nginx /assets/ cache policy. [Long-term caching without stale updates.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Hashed filenames in build output. [Standard modern web practice — confirmed.]',
        'after_evidence': '/assets/* immutable cache',
    },
    'MAF-TECH-024': {
        'decision': 'APPROVE',
        'dev_comment': 'Google Fonts loaded with display=swap and preload to reduce invisible text during load. [Text appears immediately with a fallback font, then switches to the brand font — no blank text.]',
        'next_action': 'None.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'index.html preconnect, preload, display=swap. [Fonts load without blocking page text.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Font loading strategy in index.html. [Visitors see text right away.]',
        'after_evidence': 'index.html font preload + swap',
    },
    'MAF-TECH-025': {
        'decision': 'MODIFY',
        'dev_comment': 'Route-level splitting reduces JavaScript downloaded on mobile for inner pages. React core bundle is still shared across all pages. [Mobile users get less code on sub-pages, but the base app size remains.]',
        'next_action': 'None.',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Lazy-loaded routes reduce mobile payload on navigation. [Mobile performance improved within React limits.]',
        'verify_status': 'PARTIALLY VERIFIED', 'result': 'PARTIAL',
        'verify_notes': 'Lazy routes confirmed; shared React core remains. [Better mobile loading, not identical to a lightweight static site.]',
        'after_evidence': 'React.lazy per route',
    },
    'MAF-TECH-026': {
        'decision': 'MODIFY',
        'dev_comment': 'Same React app serves mobile and desktop; Google can render JavaScript for indexing. Further HTML-in-source improvement would require SSR, which is out of scope. [Google can read our mobile site; we stay on the current technology by choice.]',
        'next_action': 'Monitor mobile indexing in GSC (Marketing).',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Mobile-first indexing supported via Google JS rendering + responsive design. [Mobile search compatibility acceptable for React SPA.]',
        'verify_status': 'PARTIALLY VERIFIED', 'result': 'PARTIAL',
        'verify_notes': 'Google renders JS; no SSR planned. [Works with Google’s mobile indexing; not ideal but accepted.]',
        'after_evidence': 'Responsive site + SEO mitigations',
    },
    'MAF-TECH-028': {
        'decision': 'FURTHER INVESTIGATION',
        'dev_comment': 'JSON-LD sameAs field is wired to dashboard Static Info social URLs. Business/marketing must enter Instagram, Facebook, LinkedIn, and YouTube profile links. [The code is ready; marketing needs to paste the company’s social media links in the admin panel.]',
        'next_action': 'Marketing: fill social profile URLs in Admin Dashboard → Static Info.',
        'impl_status': 'PENDING', 'impl_owner': 'Marketing',
        'impl_notes': 'Code supports sameAs; data entry pending. [Dev done — marketing adds the social links.]',
        'verify_status': 'PENDING EXTERNAL', 'result': 'PENDING',
        'verify_notes': 'Awaiting social URLs from business. [Google can show social icons once links are entered.]',
        'after_evidence': 'SeoProvider sameAs from Static Info API',
    },
    'MAF-TECH-030': {
        'decision': 'APPROVE',
        'dev_comment': 'Cookie consent banner implemented with Accept/Reject; Google Analytics loads only after consent. Cookie Policy page added in EN/AR. [Visitors choose whether to allow tracking cookies — required for privacy compliance.]',
        'next_action': 'Legal/marketing to review cookie policy wording if needed.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'CookieConsent.jsx + /cookie-policy pages + gated gtag. [Privacy-compliant analytics loading.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Banner and gating confirmed in SeoProvider. [Analytics only runs after user clicks Accept.]',
        'after_evidence': 'CookieConsent + conditional gtag load',
    },
    'MAF-TECH-032': {
        'decision': 'APPROVE',
        'dev_comment': 'Accessibility statement pages created at /en/accessibility and /ar/accessibility, linked in footer. [A public page explaining our accessibility commitment — often required for compliance.]',
        'next_action': 'Marketing/legal to review statement content.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'AccessibilityStatement.jsx + localized content. [Accessibility information page is live.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Pages and footer links confirmed. [Page exists; content review optional.]',
        'after_evidence': '/en/accessibility /ar/accessibility',
    },
    'MAF-TECH-040': {
        'decision': 'FURTHER INVESTIGATION',
        'dev_comment': 'Conversion tracking coded: contact form fires generate_lead; WhatsApp float button fires contact event. Marketing needs GA4 access to verify events appear in reports. [Tracking code is installed; marketing must confirm numbers show up in Google Analytics.]',
        'next_action': 'Marketing: verify generate_lead and contact events in GA4 Realtime.',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'Marketing',
        'impl_notes': 'Events in ContactForm.jsx and SiteShell.jsx; GA4 verification pending. [Dev installed tracking; marketing checks the dashboard.]',
        'verify_status': 'PENDING EXTERNAL', 'result': 'PENDING',
        'verify_notes': 'Code complete; needs GA4 login to confirm. [Someone with Analytics access must test form submission.]',
        'after_evidence': 'gtag generate_lead + contact events',
    },
    'MAF-TECH-042': {
        'decision': 'APPROVE',
        'dev_comment': 'Logo reduced to 40 KB PNG (256px) and 24 KB WebP — under 50 KB target. [Logo no longer slows pages.]',
        'next_action': 'None.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'website/public/logo-mafateeh.png. Evidence: closure-evidence/logo-size.txt',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'File size 40020 bytes < 50 KB.',
        'after_evidence': 'Compressed logo-mafateeh.png',
    },
    'MAF-TECH-043': {
        'decision': 'APPROVE',
        'dev_comment': 'New CMS uploads auto-compress via sharp (resize + WebP) before Cloudinary. Existing images use Cloudinary f_auto and srcset for responsive delivery. [Photos uploaded through the dashboard are automatically shrunk; old photos are served in optimal size.]',
        'next_action': 'Marketing: optionally re-upload very large legacy portfolio images.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'backend imageOptimize.ts + optimizedImageUrl.js on frontend. [Automatic image optimization pipeline.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Upload pipeline and srcset confirmed. [New uploads optimized; legacy images served responsively.]',
        'after_evidence': 'sharp upload + Cloudinary srcset',
    },
    'MAF-TECH-044': {
        'decision': 'APPROVE',
        'dev_comment': 'Responsive images with WebP via Cloudinary transformations (f_auto, q_auto, srcset) on blog and portfolio components. Lazy loading already present. [Images automatically adjust to screen size and modern format for faster loading.]',
        'next_action': 'None.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'optimizedImageUrl.js on BlogCard, PortfolioCard, BlogPostArticle. [Right-size images on every device.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'srcset and lazy loading in components. [Mobile gets smaller images; desktop gets full quality.]',
        'after_evidence': 'Cloudinary f_auto/srcset on image components',
    },
    'MAF-TECH-046': {
        'decision': 'APPROVE',
        'dev_comment': 'Google Fonts render-blocking reduced with preload and display=swap (same fix as MAF-TECH-024). [Page text shows immediately instead of waiting for fonts.]',
        'next_action': 'None.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Font preload strategy in index.html. [Fonts no longer block first text display.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Preload + swap confirmed. [Same as font optimization fix.]',
        'after_evidence': 'index.html font strategy',
    },
    'MAF-TECH-050': {
        'decision': 'MODIFY',
        'dev_comment': 'Accessibility statement page added. Full accessibility audit applies to the live rendered page (Lighthouse/axe), not raw HTML source — limited by React CSR architecture which we are keeping. [We added the required statement page; a full accessibility score test runs on the live site, not the code file alone.]',
        'next_action': 'Optional: run Lighthouse accessibility audit after deploy.',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Statement page live; raw-HTML a11y audit limitation accepted with React. [Compliance page done; deep audit is a separate test.]',
        'verify_status': 'PARTIALLY VERIFIED', 'result': 'PARTIAL',
        'verify_notes': 'Statement exists; full baseline needs live Lighthouse run. [Page added; formal score test can be run later.]',
        'after_evidence': 'Accessibility statement + semantic components',
    },
    'MAF-TECH-052': {
        'decision': 'MODIFY',
        'dev_comment': 'Technical debt actively reduced while staying on React: bilingual URLs, code splitting, caching, cookie consent, T&C, image pipeline, performance fixes. Remaining CSR limitations are an accepted architectural choice, not unmanaged debt. [We cleaned up many issues without rebuilding the site; some limits remain because we chose to keep React.]',
        'next_action': 'None — ongoing maintenance as normal.',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Major SEO/perf debt addressed Aug 2026. SSR debt consciously deferred. [Project is healthier; full rewrite not needed.]',
        'verify_status': 'PARTIALLY VERIFIED', 'result': 'PARTIAL',
        'verify_notes': 'Multiple fixes shipped; CSR trade-off documented. [Debt reduced significantly within chosen architecture.]',
        'after_evidence': 'Phase 1–3 SEO implementation complete',
    },
    'MAF-TECH-010': {
        'decision': 'REJECT',
        'dev_comment': 'Audit finding is outdated — dedicated contact page already exists at /en/contact and /ar/contact with working form. [The contact page was already built; the audit missed it.]',
        'next_action': 'None.',
        'impl_status': 'ALREADY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Contact page predates audit. [No work needed — page already live.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'N/A — FALSE POSITIVE',
        'verify_notes': 'Contact.jsx route confirmed. [Audit was wrong on this point.]',
        'after_evidence': '/en/contact /ar/contact live',
    },
    'MAF-TECH-031': {
        'decision': 'APPROVE',
        'dev_comment': 'T&C page + admin editor live. Approved legal content entry pending Marketing/Legal. [Page built; legal text is business task.]',
        'next_action': 'Marketing/Legal: enter T&C content',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev + Marketing',
        'impl_notes': 'TermsAndConditions model + admin UI.',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS — CONTENT PENDING',
        'verify_notes': 'Page and editor in repo.',
        'after_evidence': 'TermsAndConditions page + admin',
    },
    'MAF-TECH-033': {
        'decision': 'APPROVE',
        'dev_comment': 'favicon.ico and favicon.svg added and served with correct nginx location. [The small icon in browser tabs now displays correctly.]',
        'next_action': 'None.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'public/favicon.ico + favicon.svg. [Browser tab icon fixed.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Favicon files in public/. [Standard browser icon present.]',
        'after_evidence': 'favicon.ico + favicon.svg',
    },
    'MAF-TECH-034': {
        'decision': 'APPROVE',
        'dev_comment': 'site.webmanifest created and served with correct MIME type via nginx. [Enables “add to home screen” on mobile and proper PWA metadata.]',
        'next_action': 'None.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'public/site.webmanifest + nginx location. [Mobile install metadata available.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Manifest file confirmed. [Mobile browsers can recognize the site.]',
        'after_evidence': 'site.webmanifest',
    },
    'MAF-TECH-035': {
        'decision': 'MODIFY',
        'dev_comment': 'Common probe paths (wp-admin, admin, .env, phpmyadmin) return real HTTP 404 via nginx. Other unknown SPA URLs still return 200 with noindex — same limitation as MAF-TECH-005. [Hacker probe addresses are blocked; random invalid URLs still show a “not found” page without a server error.]',
        'next_action': 'None — accepted partial mitigation.',
        'impl_status': 'PARTIALLY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'nginx probe 404 blocks implemented. [Security scans get proper errors.]',
        'verify_status': 'PARTIALLY VERIFIED', 'result': 'PARTIAL',
        'verify_notes': 'Probe paths return 404. [Main security concern addressed.]',
        'after_evidence': 'nginx probe location blocks',
    },
    'MAF-TECH-039': {
        'decision': 'REJECT',
        'dev_comment': 'Analytics intentionally not in raw HTML — must load only after cookie consent per MAF-TECH-030 and privacy policy. Injecting gtag before consent would violate our cookie compliance. [Tracking scripts wait for user permission, which is correct for privacy law.]',
        'next_action': 'None.',
        'impl_status': 'NOT APPLICABLE', 'impl_owner': 'IT Dev',
        'impl_notes': 'Consent-gated analytics by design. [This “issue” is actually correct behavior.]',
        'verify_status': 'N/A', 'result': 'N/A — BY DESIGN',
        'verify_notes': 'gtag loads post-consent only. [Analytics after Accept is intentional.]',
        'after_evidence': 'SeoProvider conditional gtag',
    },
    'MAF-TECH-041': {
        'decision': 'REJECT',
        'dev_comment': 'Audit finding is outdated — contact form and Google Calendar booking integration already exist on the contact page. [Lead capture was already built; audit did not check current site.]',
        'next_action': 'None.',
        'impl_status': 'ALREADY RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'ContactForm.jsx + booking link on contact page. [Forms already live.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'N/A — FALSE POSITIVE',
        'verify_notes': 'Contact form confirmed in codebase. [Audit was wrong.]',
        'after_evidence': 'ContactForm + calendar booking',
    },
    'MAF-TECH-045': {
        'decision': 'APPROVE',
        'dev_comment': 'Favicon MIME types corrected: .ico served as icon, .svg as image/svg+xml — no longer PNG mislabeled as SVG. [Browser receives the icon in the correct file format.]',
        'next_action': 'None.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Proper favicon.ico + favicon.svg files. [Icon format mismatch fixed.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'Correct file types in public/. [Technical favicon issue resolved.]',
        'after_evidence': 'favicon.ico + favicon.svg correct types',
    },
    'MAF-TECH-047': {
        'decision': 'APPROVE',
        'dev_comment': 'Fonts self-hosted via @fontsource — no Google Fonts requests on public site. [Privacy and performance improved; no external font dependency.]',
        'next_action': 'None.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'main.jsx @fontsource imports; removed Google Fonts from index.html. Evidence: fonts-self-hosted.txt',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'grep confirms zero fonts.googleapis.com in public website.',
        'after_evidence': 'website/src/main.jsx @fontsource imports',
    },
    'MAF-TECH-048': {
        'decision': 'APPROVE',
        'dev_comment': 'HSTS header set in HTTPS nginx configuration as sub-item of MAF-TECH-049. [Browsers are told to always use the secure HTTPS version.]',
        'next_action': 'None.',
        'impl_status': 'RESOLVED', 'impl_owner': 'IT Dev',
        'impl_notes': 'Included in security headers deployment. [Secure connection enforced.]',
        'verify_status': 'VERIFIED (CODE)', 'result': 'PASS',
        'verify_notes': 'HSTS in deploy nginx. [Covered by MAF-TECH-049 fix.]',
        'after_evidence': 'HSTS header in nginx',
    },
}

REPAIRS = {
    'AR-001': ('COMPLETED', 'robots.txt deployed as static file. [Crawl rules file is live.]'),
    'AR-002': ('COMPLETED', 'sitemap.xml with bilingual URLs deployed. [Site map for Google is live.]'),
    'AR-003': ('COMPLETED', 'Nginx serves robots.txt and sitemap as real files, not SPA. [Server correctly delivers SEO files.]'),
    'AR-004': ('COMPLETED', 'HTTP→HTTPS and apex→www redirects in deploy nginx. [Traffic goes to correct secure URL.]'),
    'AR-005': ('PARTIALLY COMPLETED', 'Probe paths 404 + noindex on unknown SPA routes. Full whitelist needs SSR. [Partial soft-404 fix applied.]'),
    'AR-006': ('COMPLETED', 'Cache-Control + CDN-Cache-Control on assets and static files. [Caching headers active; Cloudflare UI rules pending.]'),
    'AR-007': ('COMPLETED', 'Security headers including HSTS deployed. [Security headers active.]'),
    'AR-008': ('COMPLETED', 'Canonical tags via SeoProvider + index.html. [Official URLs set per page.]'),
    'AR-009': ('COMPLETED', 'Organization JSON-LD in HTML + SeoProvider. [Company schema for Google.]'),
    'AR-010': ('COMPLETED', 'WebSite JSON-LD on homepage via SeoProvider. [Homepage schema added.]'),
    'AR-011': ('COMPLETED', 'BreadcrumbList JSON-LD per route. [Navigation breadcrumbs for search.]'),
    'AR-012': ('COMPLETED', 'OG/Twitter tags static + dynamic. [Social sharing previews configured.]'),
    'AR-013': ('COMPLETED', '/en and /ar URL routing with legacy redirects. [Bilingual URLs implemented without SSR.]'),
    'AR-014': ('COMPLETED', 'hreflang + lang/dir sync for EN/AR. [Language signals for Google.]'),
    'AR-015': ('COMPLETED', 'Vite manualChunks + React.lazy code splitting. [Faster page loads via split bundles.]'),
    'AR-016': ('CANCELLED', 'Next.js SSR rejected — business decision to keep React SPA. Mitigations applied instead. [Full rebuild not approved; partial fixes done instead.]'),
    'AR-017': ('COMPLETED', 'Font preload + display=swap in index.html. [Fonts optimized.]'),
    'AR-018': ('COMPLETED', 'Logo compression, upload pipeline sharp/WebP, Cloudinary srcset. [Image optimization complete.]'),
}

COMPLETED_DATE = '2026-08-18'
STARTED_DATE = '2026-08-11'


HML_GAP_FIXES = {
    "MAF-TECH-012": "⚠️ PARTIAL — CSR mitigated, no SSR",
    "MAF-TECH-005": "⚠️ PARTIAL — probe 404 + noindex",
    "MAF-TECH-018": "⚠️ PARTIAL — bundle reduced, measure post-deploy",
    "MAF-TECH-008": "⚠️ PARTIAL — probe 404 + noindex",
    "MAF-TECH-025": "⚠️ PARTIAL — route splitting on mobile",
    "MAF-TECH-052": "⚠️ PARTIAL — debt reduced, CSR accepted",
    "MAF-TECH-035": "⚠️ PARTIAL — probe paths 404",
    "MAF-TECH-020": "⚠️ PENDING EXTERNAL — Cloudflare UI rules",
    "MAF-TECH-027": "✅ CODE DONE — Rich Results validation pending",
    "MAF-TECH-039": "N/A — BY DESIGN (consent-gated analytics)",
    "MAF-TECH-047": "✅ CODE DONE — fonts self-hosted Aug 2026",
}

GAP_ANALYSIS_FIXES = {
    "MAF-TECH-027": (
        "✅ CODE DONE — RICH RESULTS PENDING",
        "Production Rich Results Test validation",
        "Developer: run Google Rich Results Test after deploy",
        "Rich Results Test PASS on production URL",
    ),
    "MAF-TECH-039": (
        "N/A — BY DESIGN",
        "Analytics must load after cookie consent",
        "No change — proof under MAF-TECH-040",
        "gtag loads only after Accept (SeoProvider)",
    ),
    "MAF-TECH-011": (
        "✅ CODE DONE",
        "—",
        "Routes verified in App.jsx / BlogPost.jsx",
        "Blog slug routes + redirect on missing post",
    ),
    "MAF-TECH-043": (
        "✅ CODE DONE — legacy audit optional",
        "Optional re-upload of very large legacy images",
        "Marketing: optional legacy image review",
        "Upload pipeline + Cloudinary srcset in code",
    ),
    "MAF-TECH-050": (
        "⚠️ PARTIAL",
        "Production axe/Lighthouse a11y score",
        "Run Lighthouse accessibility after deploy",
        "Accessibility statement live + optional audit",
    ),
}

REPAIR_ALT_FIXES = {
    "AR-006": (
        "PARTIALLY COMPLETED",
        "Origin cache headers COMPLETED for 019/021. MAF-TECH-020 Cloudflare dashboard rules + cache HIT proof remain PENDING EXTERNAL. [Server ready; Cloudflare login task ~5 min.]",
        "—",
    ),
    "AR-009": (
        "PARTIALLY COMPLETED",
        "Organization JSON-LD wired in HTML + SeoProvider for 027. MAF-TECH-028 sameAs awaits Marketing social URLs in Static Info. [Code ready; marketing adds links.]",
        "—",
    ),
    "AR-016": (
        "CANCELLED",
        "Next.js SSR rejected — business decision to keep React SPA. NOT used as closure evidence.",
        "012→meta/JSON-LD/sitemap | 022→splitting/WebP/fonts | 026→responsive+JS indexing | 027→SeoProvider JSON-LD | 036→/en /ar routes | 038→lang/dir | 050→a11y statement | 052→debt reduction plan",
    ),
    "AR-017": (
        "COMPLETED",
        "Font preload + display=swap for 024/046. MAF-TECH-047 resolved separately via @fontsource self-hosting (no Google Fonts requests). [Fonts now served from site bundle.]",
        "047→website/src/main.jsx @fontsource imports",
    ),
}


IMPLEMENTATION_SUMMARY = (
    "RESOLVED: 26 | PARTIALLY RESOLVED: 13 | PENDING: 2 | "
    "ALREADY RESOLVED: 2 | NOT APPLICABLE: 2"
)
VERIFICATION_SUMMARY = (
    "VERIFIED (CODE): 28 | PARTIALLY VERIFIED: 9 | PENDING EXTERNAL: 4 | "
    "NOT VERIFIED: 2 | N/A: 2"
)
DEPLOYMENT_SUMMARY = (
    "Implementation deployed; code-level verification completed where recorded. "
    "Post-deployment, production, and external verification remain pending. "
    "Final acceptance is not yet complete."
)

HML_GAP_FIXES = {
    "MAF-TECH-012": "⚠️ PARTIAL — CSR mitigated, no SSR",
    "MAF-TECH-005": "⚠️ PARTIAL — probe 404 + noindex",
    "MAF-TECH-018": "⚠️ PARTIAL — bundle reduced, measure post-deploy",
    "MAF-TECH-008": "⚠️ PARTIAL — probe 404 + noindex",
    "MAF-TECH-025": "⚠️ PARTIAL — route splitting on mobile",
    "MAF-TECH-052": "⚠️ PARTIAL — debt reduced, CSR accepted",
    "MAF-TECH-035": "⚠️ PARTIAL — probe paths 404",
    "MAF-TECH-020": "⚠️ PENDING EXTERNAL — Cloudflare UI rules",
    "MAF-TECH-027": "✅ CODE DONE — Rich Results validation pending",
    "MAF-TECH-039": "N/A — BY DESIGN (consent-gated analytics)",
    "MAF-TECH-047": "✅ CODE DONE — fonts self-hosted Aug 2026",
}

GAP_ANALYSIS_FIXES = {
    "MAF-TECH-027": (
        "✅ CODE DONE — RICH RESULTS PENDING",
        "Production Rich Results Test validation",
        "Developer: run Google Rich Results Test after deploy",
        "Rich Results Test PASS on production URL",
    ),
    "MAF-TECH-039": (
        "N/A — BY DESIGN",
        "Analytics must load after cookie consent",
        "No change — proof under MAF-TECH-040",
        "gtag loads only after Accept (SeoProvider)",
    ),
    "MAF-TECH-011": (
        "✅ CODE DONE",
        "—",
        "Routes verified in App.jsx / BlogPost.jsx",
        "Blog slug routes + redirect on missing post",
    ),
    "MAF-TECH-043": (
        "✅ CODE DONE — legacy audit optional",
        "Optional re-upload of very large legacy images",
        "Marketing: optional legacy image review",
        "Upload pipeline + Cloudinary srcset in code",
    ),
    "MAF-TECH-050": (
        "⚠️ PARTIAL",
        "Production axe/Lighthouse a11y score",
        "Run Lighthouse accessibility after deploy",
        "Accessibility statement live + optional audit",
    ),
}

REPAIR_ALT_FIXES = {
    "AR-006": (
        "PARTIALLY COMPLETED",
        "Origin cache headers COMPLETED for 019/021. MAF-TECH-020 Cloudflare dashboard rules + cache HIT proof remain PENDING EXTERNAL. [Server ready; Cloudflare login task ~5 min.]",
        "—",
    ),
    "AR-009": (
        "PARTIALLY COMPLETED",
        "Organization JSON-LD wired in HTML + SeoProvider for 027. MAF-TECH-028 sameAs awaits Marketing social URLs in Static Info. [Code ready; marketing adds links.]",
        "—",
    ),
    "AR-016": (
        "CANCELLED",
        "Next.js SSR rejected — business decision to keep React SPA. NOT used as closure evidence.",
        "012→meta/JSON-LD/sitemap | 022→splitting/WebP/fonts | 026→responsive+JS indexing | 027→SeoProvider JSON-LD | 036→/en /ar routes | 038→lang/dir | 050→a11y statement | 052→debt reduction plan",
    ),
    "AR-017": (
        "COMPLETED",
        "Font preload + display=swap for 024/046. MAF-TECH-047 resolved separately via @fontsource self-hosting (no Google Fonts requests). [Fonts now served from site bundle.]",
        "047→website/src/main.jsx @fontsource imports",
    ),
}



def main():
    wb = openpyxl.load_workbook(WORKBOOK)
    # Sheet 3: Dev Decisions
    ws3 = wb['3. Dev Decisions']
    for r in range(2, ws3.max_row + 1):
        issue_id = ws3.cell(r, 1).value
        if issue_id in ISSUES:
            d = ISSUES[issue_id]
            ws3.cell(r, 4, decision_cell(d['decision']))
            ws3.cell(r, 5, d['dev_comment'])
            ws3.cell(r, 6, COMPLETED_DATE)
            ws3.cell(r, 7, d['next_action'])

    # Sheet 4: Impl Tracker
    ws4 = wb['4. Impl Tracker']
    for r in range(2, ws4.max_row + 1):
        issue_id = ws4.cell(r, 1).value
        if issue_id in ISSUES:
            d = ISSUES[issue_id]
            ws4.cell(r, 5, d['impl_status'])
            ws4.cell(r, 6, d['impl_owner'])
            ws4.cell(r, 7, COMPLETED_DATE)
            ws4.cell(r, 8, STARTED_DATE)
            if d['impl_status'] in ('RESOLVED', 'ALREADY RESOLVED', 'NOT APPLICABLE'):
                ws4.cell(r, 9, COMPLETED_DATE)
            elif d['impl_status'] == 'PARTIALLY RESOLVED':
                ws4.cell(r, 9, COMPLETED_DATE)
            ws4.cell(r, 10, d['impl_notes'])

    # Sheet 5: Verify Tracker
    ws5 = wb['5. Verify Tracker']
    for r in range(2, ws5.max_row + 1):
        issue_id = ws5.cell(r, 1).value
        if issue_id in ISSUES:
            d = ISSUES[issue_id]
            ws5.cell(r, 4, d['verify_status'])
            ws5.cell(r, 6, d.get('after_evidence', ''))
            ws5.cell(r, 7, 'IT Dev Team')
            if d['verify_status'] not in ('PENDING EXTERNAL', 'N/A', 'NOT VERIFIED'):
                ws5.cell(r, 8, COMPLETED_DATE)
            ws5.cell(r, 9, d['result'])
            ws5.cell(r, 10, d['verify_notes'])

    ws2 = wb["2. HML Matrix"]
    for r in range(2, ws2.max_row + 1):
        iid = ws2.cell(r, 2).value
        if iid in HML_GAP_FIXES:
            ws2.cell(r, 8, HML_GAP_FIXES[iid])

    ws6 = wb["6. Gap Analysis"]
    for r in range(2, ws6.max_row + 1):
        iid = ws6.cell(r, 1).value
        if iid in GAP_ANALYSIS_FIXES:
            gt, miss, act, vr = GAP_ANALYSIS_FIXES[iid]
            ws6.cell(r, 4, gt); ws6.cell(r, 5, miss); ws6.cell(r, 6, act); ws6.cell(r, 7, vr)

    ws7 = wb["7. Repair Manifest"]
    ws7.cell(1, 8, "Notes"); ws7.cell(1, 9, "Alternative Artifacts")
    for r in range(2, ws7.max_row + 1):
        rid = ws7.cell(r, 1).value
        if rid in REPAIRS:
            ws7.cell(r, 7, REPAIRS[rid][0]); ws7.cell(r, 8, REPAIRS[rid][1])
        if rid in REPAIR_ALT_FIXES:
            st, note, alt = REPAIR_ALT_FIXES[rid]
            ws7.cell(r, 7, st); ws7.cell(r, 8, note); ws7.cell(r, 9, alt)

    ws8 = wb["8. Summary"]
    for r in range(1, ws8.max_row + 1):
        m = ws8.cell(r, 1).value
        if m == "Implementation Status": ws8.cell(r, 2, IMPLEMENTATION_SUMMARY)
        elif m == "Verification Status": ws8.cell(r, 2, VERIFICATION_SUMMARY)
        elif m in ("Production / Deployment Status", "Production Changes", "No Production Changes"):
            ws8.cell(r, 1, "Production / Deployment Status"); ws8.cell(r, 2, DEPLOYMENT_SUMMARY)

    changelog = json.loads((ROOT / "_changelog.json").read_text(encoding="utf-8"))
    if "9. Change Log" in wb.sheetnames: del wb["9. Change Log"]
    ws9 = wb.create_sheet("9. Change Log")
    ws9.append(("Sheet", "Issue / Artifact", "Field", "Old Value", "New Value", "Reason", "Evidence Reference"))
    for row in changelog: ws9.append(list(row))

    wb.save(WORKBOOK)
    verify_counts = {}
    for d in ISSUES.values():
        vs = d["verify_status"]; verify_counts[vs] = verify_counts.get(vs, 0) + 1
    print("Saved:", WORKBOOK)
    print("Implementation summary:", IMPLEMENTATION_SUMMARY)
    print("Verification summary:", VERIFICATION_SUMMARY)
    print("Verify counts:", verify_counts)
    return verify_counts

if __name__ == "__main__":
    main()
